"""
Script to create a beautiful Gantt chart in Excel
Tạo Gantt chart đẹp giống hình mẫu với màu sắc và formatting chuyên nghiệp
"""

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# Định nghĩa dữ liệu project
project_data = [
    # Phase 1: Research & Requirements Analysis
    {"phase": "Phase 1: Research", "task": "Literature Review", "start": "2025-03-01", "duration": 14, "color": "FFD699"},
    {"phase": "Phase 1: Research", "task": "Competitive Analysis", "start": "2025-03-08", "duration": 14, "color": "FFD699"},
    {"phase": "Phase 1: Research", "task": "Requirements Gathering", "start": "2025-03-15", "duration": 14, "color": "FFD699"},
    {"phase": "Phase 1: Research", "task": "Use Case Development", "start": "2025-03-22", "duration": 21, "color": "FFD699"},

    # Phase 2: System Design
    {"phase": "Phase 2: Design", "task": "Architecture Design", "start": "2025-04-13", "duration": 14, "color": "99CCFF"},
    {"phase": "Phase 2: Design", "task": "Database Schema Design", "start": "2025-04-20", "duration": 21, "color": "99CCFF"},
    {"phase": "Phase 2: Design", "task": "UI/UX Design", "start": "2025-05-04", "duration": 21, "color": "99CCFF"},
    {"phase": "Phase 2: Design", "task": "API Endpoint Design", "start": "2025-05-18", "duration": 21, "color": "99CCFF"},
    {"phase": "Phase 2: Design", "task": "Integration Planning", "start": "2025-05-25", "duration": 14, "color": "99CCFF"},

    # Phase 3: Backend Implementation
    {"phase": "Phase 3: Backend", "task": "Dev Environment Setup", "start": "2025-06-08", "duration": 7, "color": "CC99FF"},
    {"phase": "Phase 3: Backend", "task": "Authentication System", "start": "2025-06-15", "duration": 14, "color": "CC99FF"},
    {"phase": "Phase 3: Backend", "task": "Core Models Development", "start": "2025-06-22", "duration": 21, "color": "CC99FF"},
    {"phase": "Phase 3: Backend", "task": "Business Logic Controllers", "start": "2025-07-06", "duration": 21, "color": "CC99FF"},
    {"phase": "Phase 3: Backend", "task": "Advanced Features", "start": "2025-07-20", "duration": 21, "color": "CC99FF"},
    {"phase": "Phase 3: Backend", "task": "Payment Integration", "start": "2025-08-03", "duration": 21, "color": "CC99FF"},
    {"phase": "Phase 3: Backend", "task": "Community Features", "start": "2025-08-10", "duration": 14, "color": "CC99FF"},
    {"phase": "Phase 3: Backend", "task": "Media Upload Integration", "start": "2025-08-17", "duration": 7, "color": "CC99FF"},

    # Phase 4: Frontend Development
    {"phase": "Phase 4: Frontend", "task": "Landing & Auth UI", "start": "2025-08-24", "duration": 14, "color": "99FF99"},
    {"phase": "Phase 4: Frontend", "task": "Course Catalog & Details", "start": "2025-09-07", "duration": 14, "color": "99FF99"},
    {"phase": "Phase 4: Frontend", "task": "Dashboard Interfaces", "start": "2025-09-14", "duration": 21, "color": "99FF99"},
    {"phase": "Phase 4: Frontend", "task": "Lesson & Quiz Interfaces", "start": "2025-09-28", "duration": 21, "color": "99FF99"},
    {"phase": "Phase 4: Frontend", "task": "Payment & E-commerce", "start": "2025-10-12", "duration": 14, "color": "99FF99"},
    {"phase": "Phase 4: Frontend", "task": "Community & Profile", "start": "2025-10-19", "duration": 14, "color": "99FF99"},
    {"phase": "Phase 4: Frontend", "task": "UI/UX Polish", "start": "2025-10-26", "duration": 14, "color": "99FF99"},

    # Phase 5: Testing & QA
    {"phase": "Phase 5: Testing", "task": "Unit Testing", "start": "2025-10-05", "duration": 21, "color": "FFFF99"},
    {"phase": "Phase 5: Testing", "task": "Integration Testing", "start": "2025-10-19", "duration": 14, "color": "FFFF99"},
    {"phase": "Phase 5: Testing", "task": "Security Testing", "start": "2025-10-26", "duration": 14, "color": "FFFF99"},
    {"phase": "Phase 5: Testing", "task": "Performance Testing", "start": "2025-11-02", "duration": 14, "color": "FFFF99"},
    {"phase": "Phase 5: Testing", "task": "Cross-browser Testing", "start": "2025-11-09", "duration": 14, "color": "FFFF99"},
    {"phase": "Phase 5: Testing", "task": "User Acceptance Testing", "start": "2025-11-16", "duration": 7, "color": "FFFF99"},

    # Phase 6: MVC Refactoring
    {"phase": "Phase 6: Refactoring", "task": "Create Firebase Helpers", "start": "2025-10-19", "duration": 7, "color": "FF9999"},
    {"phase": "Phase 6: Refactoring", "task": "Build Missing Models", "start": "2025-10-26", "duration": 7, "color": "FF9999"},
    {"phase": "Phase 6: Refactoring", "task": "Enhance Existing Models", "start": "2025-10-26", "duration": 14, "color": "FF9999"},
    {"phase": "Phase 6: Refactoring", "task": "Fix N+1 Query Problems", "start": "2025-11-02", "duration": 7, "color": "FF9999"},
    {"phase": "Phase 6: Refactoring", "task": "Refactor All Controllers", "start": "2025-11-09", "duration": 7, "color": "FF9999"},
    {"phase": "Phase 6: Refactoring", "task": "Fix Database Issues", "start": "2025-11-16", "duration": 3, "color": "FF9999"},
    {"phase": "Phase 6: Refactoring", "task": "Final Verification", "start": "2025-11-20", "duration": 3, "color": "FF9999"},

    # Phase 7: Deployment & Docs
    {"phase": "Phase 7: Deployment", "task": "Production Deployment", "start": "2025-11-09", "duration": 7, "color": "99FFFF"},
    {"phase": "Phase 7: Deployment", "task": "Academic Report Writing", "start": "2025-11-16", "duration": 7, "color": "99FFFF"},
    {"phase": "Phase 7: Deployment", "task": "Technical Documentation", "start": "2025-11-23", "duration": 3, "color": "99FFFF"},
    {"phase": "Phase 7: Deployment", "task": "Presentation Preparation", "start": "2025-11-23", "duration": 7, "color": "99FFFF"},
    {"phase": "Phase 7: Deployment", "task": "Final Submission", "start": "2025-11-29", "duration": 1, "color": "99FFFF"},
]

# Tạo DataFrame
df = pd.DataFrame(project_data)

# Chuyển đổi string sang datetime
df['start'] = pd.to_datetime(df['start'])
df['end'] = df['start'] + pd.to_timedelta(df['duration'], unit='D')

# Tính start offset (ngày từ project start)
project_start = df['start'].min()
df['start_offset'] = (df['start'] - project_start).dt.days

# Tạo Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Gantt Chart"

# Định nghĩa styles
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
cell_font = Font(name='Calibri', size=10)
alignment_center = Alignment(horizontal='center', vertical='center')
alignment_left = Alignment(horizontal='left', vertical='center')
border_thin = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

# Headers
headers = ['Phase', 'Task Name', 'Start Date', 'End Date', 'Duration (Days)', 'Start Offset', 'Duration']
ws.append(headers)

# Format header row
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment_center
    cell.border = border_thin

# Populate data
for idx, row in df.iterrows():
    ws.append([
        row['phase'],
        row['task'],
        row['start'].strftime('%Y-%m-%d'),
        row['end'].strftime('%Y-%m-%d'),
        row['duration'],
        row['start_offset'],  # For chart calculation
        row['duration']        # For chart display
    ])

# Format data rows
for row_num in range(2, len(df) + 2):
    for col_num in range(1, 8):
        cell = ws.cell(row=row_num, column=col_num)
        cell.font = cell_font
        cell.border = border_thin

        if col_num in [3, 4]:  # Date columns
            cell.alignment = alignment_center
        elif col_num in [5, 6, 7]:  # Number columns
            cell.alignment = alignment_center
        else:
            cell.alignment = alignment_left

# Set column widths
ws.column_dimensions['A'].width = 20  # Phase
ws.column_dimensions['B'].width = 30  # Task Name
ws.column_dimensions['C'].width = 12  # Start Date
ws.column_dimensions['D'].width = 12  # End Date
ws.column_dimensions['E'].width = 15  # Duration
ws.column_dimensions['F'].width = 12  # Start Offset
ws.column_dimensions['G'].width = 12  # Duration

# Apply color coding to Phase column
phase_colors = {
    "Phase 1: Research": "FFD699",
    "Phase 2: Design": "99CCFF",
    "Phase 3: Backend": "CC99FF",
    "Phase 4: Frontend": "99FF99",
    "Phase 5: Testing": "FFFF99",
    "Phase 6: Refactoring": "FF9999",
    "Phase 7: Deployment": "99FFFF"
}

for row_num in range(2, len(df) + 2):
    phase_cell = ws.cell(row=row_num, column=1)
    phase = phase_cell.value
    if phase in phase_colors:
        phase_cell.fill = PatternFill(start_color=phase_colors[phase],
                                       end_color=phase_colors[phase],
                                       fill_type='solid')

# Create Stacked Bar Chart
chart = BarChart()
chart.type = "bar"
chart.style = 12
chart.title = "UniLearn/EduLearn Project Timeline (March - November 2025)"
chart.y_axis.title = "Tasks"
chart.x_axis.title = "Days from Project Start"

# Data for chart
# Start Offset (invisible) + Duration (visible)
data = Reference(ws, min_col=6, min_row=1, max_row=len(df) + 1, max_col=7)
cats = Reference(ws, min_col=2, min_row=2, max_row=len(df) + 1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# Format chart
chart.width = 20  # Width in inches
chart.height = 15  # Height in inches
chart.legend = None  # Remove legend for cleaner look

# Style the chart series
# Series 0 = Start Offset (make invisible)
# Series 1 = Duration (make colorful)
chart.series[0].graphicalProperties.noFill = True  # Hide start offset
chart.series[0].graphicalProperties.line.noFill = True

# Apply colors to Duration bars based on phase
# Note: openpyxl has limitations with per-bar coloring in stacked charts
# This will apply uniform color - for per-bar colors, consider using a different library

# Position chart
ws.add_chart(chart, "I2")

# Add a legend worksheet
legend_ws = wb.create_sheet("Legend")
legend_ws.column_dimensions['A'].width = 25
legend_ws.column_dimensions['B'].width = 15

legend_data = [
    ["Phase", "Color"],
    ["Phase 1: Research & Requirements", "Orange/Peach"],
    ["Phase 2: System Design", "Light Blue"],
    ["Phase 3: Backend Implementation", "Light Purple"],
    ["Phase 4: Frontend Development", "Light Green"],
    ["Phase 5: Testing & QA", "Light Yellow"],
    ["Phase 6: MVC Refactoring", "Light Red"],
    ["Phase 7: Deployment & Docs", "Light Cyan"]
]

for row_idx, (phase, color_desc) in enumerate(legend_data, 1):
    legend_ws.cell(row=row_idx, column=1, value=phase)
    legend_ws.cell(row=row_idx, column=2, value=color_desc)

    if row_idx == 1:  # Header
        legend_ws.cell(row=row_idx, column=1).font = header_font
        legend_ws.cell(row=row_idx, column=2).font = header_font
        legend_ws.cell(row=row_idx, column=1).fill = header_fill
        legend_ws.cell(row=row_idx, column=2).fill = header_fill
    else:  # Data rows
        phase_name = list(phase_colors.keys())[row_idx - 2]
        color_code = phase_colors[phase_name]
        legend_ws.cell(row=row_idx, column=2).fill = PatternFill(
            start_color=color_code,
            end_color=color_code,
            fill_type='solid'
        )

# Add project info worksheet
info_ws = wb.create_sheet("Project Info")
info_ws.column_dimensions['A'].width = 30
info_ws.column_dimensions['B'].width = 50

project_info = [
    ["Project Name", "UniLearn/EduLearn - Learning Management System"],
    ["Student ID", "GCS220124"],
    ["Greenwich ID", "001322934"],
    ["Project Duration", "40 weeks (280 days)"],
    ["Start Date", "March 1, 2025"],
    ["End Date", "November 29, 2025"],
    ["Total Phases", "7"],
    ["Total Tasks", str(len(df))],
    ["Live URL", "https://x.huy.global/"],
    ["GitHub", "https://github.com/givhvy/FINAL-PROJECT"],
    ["", ""],
    ["Phase Breakdown", ""],
    ["Phase 1: Research & Requirements", "6 weeks (March 1 - April 12)"],
    ["Phase 2: System Design", "8 weeks (April 13 - June 7)"],
    ["Phase 3: Backend Implementation", "11 weeks (June 8 - August 23)"],
    ["Phase 4: Frontend Development", "11 weeks (August 24 - November 8)"],
    ["Phase 5: Testing & QA", "7 weeks (October 5 - November 22)"],
    ["Phase 6: MVC Refactoring", "5 weeks (October 19 - November 22)"],
    ["Phase 7: Deployment & Docs", "3 weeks (November 9 - November 29)"],
]

for row_idx, (key, value) in enumerate(project_info, 1):
    info_ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
    info_ws.cell(row=row_idx, column=2, value=value)

# Save workbook
output_file = "UniLearn_Gantt_Chart_Beautiful.xlsx"
wb.save(output_file)

print(f"[SUCCESS] Gantt chart created successfully: {output_file}")
print(f"[INFO] Total tasks: {len(df)}")
print(f"[INFO] Project duration: {(df['end'].max() - df['start'].min()).days} days")
print(f"[INFO] 7 phases with color coding")
print(f"[INFO] Professional Excel format with chart")
print(f"\nOpen the file in Excel to view the beautiful Gantt chart!")
