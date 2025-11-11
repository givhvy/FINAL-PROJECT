"""
Advanced Gantt Chart Creator with Timeline Visualization
Tạo Gantt chart với timeline visualization đẹp như hình mẫu
"""

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Project data
project_data = [
    # Phase 1
    {"phase": "Phase 1", "task": "Literature Review", "start": "2025-03-01", "weeks": 2},
    {"phase": "Phase 1", "task": "Competitive Analysis", "start": "2025-03-08", "weeks": 2},
    {"phase": "Phase 1", "task": "Requirements Gathering", "start": "2025-03-15", "weeks": 2},
    {"phase": "Phase 1", "task": "Use Case Development", "start": "2025-03-22", "weeks": 3},
    # Phase 2
    {"phase": "Phase 2", "task": "Architecture Design", "start": "2025-04-13", "weeks": 2},
    {"phase": "Phase 2", "task": "Database Schema Design", "start": "2025-04-20", "weeks": 3},
    {"phase": "Phase 2", "task": "UI/UX Design", "start": "2025-05-04", "weeks": 3},
    {"phase": "Phase 2", "task": "API Endpoint Design", "start": "2025-05-18", "weeks": 3},
    {"phase": "Phase 2", "task": "Integration Planning", "start": "2025-05-25", "weeks": 2},
    # Phase 3
    {"phase": "Phase 3", "task": "Dev Environment Setup", "start": "2025-06-08", "weeks": 1},
    {"phase": "Phase 3", "task": "Authentication System", "start": "2025-06-15", "weeks": 2},
    {"phase": "Phase 3", "task": "Core Models Development", "start": "2025-06-22", "weeks": 3},
    {"phase": "Phase 3", "task": "Business Logic Controllers", "start": "2025-07-06", "weeks": 3},
    {"phase": "Phase 3", "task": "Advanced Features", "start": "2025-07-20", "weeks": 3},
    {"phase": "Phase 3", "task": "Payment Integration", "start": "2025-08-03", "weeks": 3},
    {"phase": "Phase 3", "task": "Community Features", "start": "2025-08-10", "weeks": 2},
    {"phase": "Phase 3", "task": "Media Upload Integration", "start": "2025-08-17", "weeks": 1},
    # Phase 4
    {"phase": "Phase 4", "task": "Landing & Auth UI", "start": "2025-08-24", "weeks": 2},
    {"phase": "Phase 4", "task": "Course Catalog & Details", "start": "2025-09-07", "weeks": 2},
    {"phase": "Phase 4", "task": "Dashboard Interfaces", "start": "2025-09-14", "weeks": 3},
    {"phase": "Phase 4", "task": "Lesson & Quiz Interfaces", "start": "2025-09-28", "weeks": 3},
    {"phase": "Phase 4", "task": "Payment & E-commerce", "start": "2025-10-12", "weeks": 2},
    {"phase": "Phase 4", "task": "Community & Profile", "start": "2025-10-19", "weeks": 2},
    {"phase": "Phase 4", "task": "UI/UX Polish", "start": "2025-10-26", "weeks": 2},
    # Phase 5
    {"phase": "Phase 5", "task": "Unit Testing", "start": "2025-10-05", "weeks": 3},
    {"phase": "Phase 5", "task": "Integration Testing", "start": "2025-10-19", "weeks": 2},
    {"phase": "Phase 5", "task": "Security Testing", "start": "2025-10-26", "weeks": 2},
    {"phase": "Phase 5", "task": "Performance Testing", "start": "2025-11-02", "weeks": 2},
    {"phase": "Phase 5", "task": "Cross-browser Testing", "start": "2025-11-09", "weeks": 2},
    {"phase": "Phase 5", "task": "User Acceptance Testing", "start": "2025-11-16", "weeks": 1},
    # Phase 6
    {"phase": "Phase 6", "task": "Create Firebase Helpers", "start": "2025-10-19", "weeks": 1},
    {"phase": "Phase 6", "task": "Build Missing Models", "start": "2025-10-26", "weeks": 1},
    {"phase": "Phase 6", "task": "Enhance Existing Models", "start": "2025-10-26", "weeks": 2},
    {"phase": "Phase 6", "task": "Fix N+1 Query Problems", "start": "2025-11-02", "weeks": 1},
    {"phase": "Phase 6", "task": "Refactor All Controllers", "start": "2025-11-09", "weeks": 1},
    {"phase": "Phase 6", "task": "Fix Database Issues", "start": "2025-11-16", "weeks": 0.5},
    {"phase": "Phase 6", "task": "Final Verification", "start": "2025-11-20", "weeks": 0.5},
    # Phase 7
    {"phase": "Phase 7", "task": "Production Deployment", "start": "2025-11-09", "weeks": 1},
    {"phase": "Phase 7", "task": "Academic Report Writing", "start": "2025-11-16", "weeks": 1},
    {"phase": "Phase 7", "task": "Technical Documentation", "start": "2025-11-23", "weeks": 0.5},
    {"phase": "Phase 7", "task": "Presentation Preparation", "start": "2025-11-23", "weeks": 1},
    {"phase": "Phase 7", "task": "Final Submission", "start": "2025-11-29", "weeks": 0.2},
]

# Phase colors
phase_colors = {
    "Phase 1": "FFD699",  # Orange pastel
    "Phase 2": "99CCFF",  # Blue pastel
    "Phase 3": "CC99FF",  # Purple pastel
    "Phase 4": "99FF99",  # Green pastel
    "Phase 5": "FFFF99",  # Yellow pastel
    "Phase 6": "FF9999",  # Red pastel
    "Phase 7": "99FFFF",  # Cyan pastel
}

# Create DataFrame
df = pd.DataFrame(project_data)
df['start'] = pd.to_datetime(df['start'])
df['duration_days'] = (df['weeks'] * 7).astype(int)
df['end'] = df['start'] + pd.to_timedelta(df['duration_days'], unit='D')

# Calculate project timeline
project_start = df['start'].min()
project_end = df['end'].max()
total_weeks = int((project_end - project_start).days / 7) + 1

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Gantt Timeline"

# Styles
header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
cell_font = Font(name='Calibri', size=9)
border_thin = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Set up headers
ws['A1'] = 'Phase'
ws['B1'] = 'Task Name'
ws['C1'] = 'Start'
ws['D1'] = 'Weeks'

# Create week headers (W1, W2, ..., W40)
for week_num in range(1, total_weeks + 1):
    col = 4 + week_num  # Start from column E (5th column)
    cell = ws.cell(row=1, column=col)
    cell.value = f'W{week_num}'
    cell.font = Font(name='Calibri', size=8, bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
    ws.column_dimensions[get_column_letter(col)].width = 2.5

# Format first 4 columns headers
for col in range(1, 5):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_thin

# Set column widths
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 6

# Populate data and create timeline bars
for idx, row in df.iterrows():
    row_num = idx + 2

    # Phase
    ws.cell(row=row_num, column=1, value=row['phase']).font = cell_font

    # Task name
    task_cell = ws.cell(row=row_num, column=2, value=row['task'])
    task_cell.font = cell_font
    task_cell.alignment = Alignment(horizontal='left', vertical='center')

    # Start date
    ws.cell(row=row_num, column=3, value=row['start'].strftime('%m/%d')).font = cell_font

    # Weeks
    ws.cell(row=row_num, column=4, value=f"{row['weeks']:.1f}").font = cell_font

    # Calculate which weeks this task spans
    task_start_week = int((row['start'] - project_start).days / 7) + 1
    task_duration_weeks = int(row['weeks'])

    # Color the appropriate week cells
    phase_color = phase_colors[row['phase']]
    fill = PatternFill(start_color=phase_color, end_color=phase_color, fill_type='solid')

    for week_offset in range(task_duration_weeks + 1):
        week_num = task_start_week + week_offset
        if week_num <= total_weeks:
            col = 4 + week_num
            cell = ws.cell(row=row_num, column=col)
            cell.fill = fill
            cell.border = border_thin

# Apply borders to all data cells
for row_num in range(2, len(df) + 2):
    for col_num in range(1, 5 + total_weeks):
        cell = ws.cell(row=row_num, column=col_num)
        cell.border = border_thin
        if col_num <= 4:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Freeze panes (freeze first 4 columns and first row)
ws.freeze_panes = 'E2'

# Add month markers
# Create a row above headers for months
ws.insert_rows(1)
ws.row_dimensions[1].height = 20

# Calculate months
current_date = project_start
month_starts = {}
week = 1

while current_date <= project_end:
    month_year = current_date.strftime('%b %Y')
    if month_year not in month_starts:
        month_starts[month_year] = week
    current_date += timedelta(days=7)
    week += 1

# Merge cells for months and add labels
for month_label, start_week in month_starts.items():
    # Find how many weeks this month spans
    weeks_in_month = 4  # Approximate
    start_col = 4 + start_week
    end_col = start_col + weeks_in_month - 1

    if end_col > 4 + total_weeks:
        end_col = 4 + total_weeks

    # Merge cells
    if start_col <= end_col:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        month_cell = ws.cell(row=1, column=start_col)
        month_cell.value = month_label
        month_cell.font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
        month_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        month_cell.alignment = Alignment(horizontal='center', vertical='center')
        month_cell.border = border_thin

# Add title
ws.insert_rows(1)
ws.merge_cells('A1:G1')
title_cell = ws['A1']
title_cell.value = 'UniLearn/EduLearn Project Timeline - Gantt Chart'
title_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
title_cell.fill = PatternFill(start_color='203764', end_color='203764', fill_type='solid')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 25

# Add legend sheet
legend_ws = wb.create_sheet("Legend")
legend_ws.column_dimensions['A'].width = 30
legend_ws.column_dimensions['B'].width = 30

legend_title = [
    ["UniLearn/EduLearn Project", ""],
    ["Project Timeline Legend", ""],
    ["", ""],
    ["Phase", "Description"],
]

phase_descriptions = [
    ["Phase 1", "Research & Requirements Analysis (6 weeks)"],
    ["Phase 2", "System Design (8 weeks)"],
    ["Phase 3", "Backend Implementation (11 weeks)"],
    ["Phase 4", "Frontend Development (11 weeks)"],
    ["Phase 5", "Testing & QA (7 weeks)"],
    ["Phase 6", "MVC Refactoring (5 weeks)"],
    ["Phase 7", "Deployment & Documentation (3 weeks)"],
]

# Write legend
for row_idx, (key, value) in enumerate(legend_title, 1):
    legend_ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True, size=12)
    legend_ws.cell(row=row_idx, column=2, value=value)

start_row = len(legend_title) + 1
for row_idx, (phase, desc) in enumerate(phase_descriptions, start_row):
    phase_cell = legend_ws.cell(row=row_idx, column=1, value=phase)
    phase_cell.font = Font(bold=True)
    phase_cell.fill = PatternFill(
        start_color=phase_colors[phase],
        end_color=phase_colors[phase],
        fill_type='solid'
    )
    legend_ws.cell(row=row_idx, column=2, value=desc)

# Save
output_file = "UniLearn_Gantt_Chart_Timeline.xlsx"
wb.save(output_file)

print(f"[SUCCESS] Advanced Gantt chart created: {output_file}")
print(f"[INFO] Total tasks: {len(df)}")
print(f"[INFO] Total weeks: {total_weeks}")
print(f"[INFO] Timeline visualization with week-by-week coloring")
print(f"[INFO] Month markers included")
print(f"\nThis file shows a beautiful timeline view like your sample image!")
